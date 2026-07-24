"""Parse EEX primary-market auction reports into the monitor's history CSV.

Handles both report generations: legacy .xls (2012-2019, Excel serial
dates, explicit Cover Ratio column, no Status column) and modern .xlsx
(2020+, ISO dates, Status column). Only EUA spot contracts (T2PA/T3PA)
count — aviation (EUAA) and cancelled auctions are dropped. Days with
several auctions (EU + DE + PL) are volume-weighted into one daily row.

One-off archive import:
    python -m eua_monitor.eex_import <zip-or-directory>
writes data/eua_history.csv. The same parser serves fetch.py for the
live current-year report.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import sys
import warnings
import zipfile
from pathlib import Path

from . import config

EUA_CONTRACTS = {"T2PA", "T3PA"}

_HEADER_KEYS = {
    "date": ("Date",),
    "name": ("Auction Name",),
    "contract": ("Contract",),
    "status": ("Status",),
    "price": ("Auction Price",),
    "volume": ("Auction Volume",),
    "bids": ("Total Amount of Bids",),
    "cover": ("Cover Ratio",),
}


def _map_header(cells: list) -> dict[str, int] | None:
    text = [str(c).strip() if c is not None else "" for c in cells]
    cols = {}
    for key, prefixes in _HEADER_KEYS.items():
        for i, cell in enumerate(text):
            if any(cell.startswith(p) for p in prefixes):
                cols[key] = i
                break
    return cols if {"date", "contract", "price", "volume"} <= cols.keys() else None


def _rows_from_grid(grid, to_date) -> list[dict]:
    cols = None
    out = []
    for raw in grid:
        if cols is None:
            cols = _map_header(list(raw))
            continue
        def cell(key):
            i = cols.get(key)
            return raw[i] if i is not None and i < len(raw) else None
        contract = str(cell("contract") or "").strip()
        price, volume = cell("price"), cell("volume")
        if contract not in EUA_CONTRACTS or price is None or volume in (None, 0, ""):
            continue
        status = str(cell("status") or "successful").strip().lower()
        if status != "successful":
            continue
        day = to_date(cell("date"))
        if day is None:
            continue
        bids = cell("bids")
        cover = cell("cover")
        out.append(
            {
                "date": day,
                "price": float(price),
                "volume": float(volume),
                "bids": float(bids) if bids not in (None, "") else None,
                "cover": float(cover) if cover not in (None, "") else None,
            }
        )
    return out


def parse_xlsx_bytes(data: bytes) -> list[dict]:
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    ws = wb[wb.sheetnames[0]]

    def to_date(v):
        if isinstance(v, dt.datetime):
            return v.date().isoformat()
        if isinstance(v, dt.date):
            return v.isoformat()
        return str(v)[:10] if v else None

    return _rows_from_grid(ws.iter_rows(values_only=True), to_date)


def parse_xls_bytes(data: bytes) -> list[dict]:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)

    def to_date(v):
        if isinstance(v, float) and v > 20000:  # Excel serial
            return dt.datetime(*xlrd.xldate_as_tuple(v, book.datemode)[:3]).date().isoformat()
        return str(v)[:10] if v else None

    return _rows_from_grid(
        (sheet.row_values(i) for i in range(sheet.nrows)), to_date
    )


def parse_report(name: str, data: bytes) -> list[dict]:
    return parse_xls_bytes(data) if name.endswith(".xls") else parse_xlsx_bytes(data)


def aggregate_daily(rows: list[dict]) -> list[dict]:
    """Volume-weighted daily price; bids summed; cover = bids / volume."""
    by_day: dict[str, list[dict]] = {}
    for row in rows:
        by_day.setdefault(row["date"], []).append(row)
    out = []
    for day in sorted(by_day):
        group = by_day[day]
        vol = sum(r["volume"] for r in group)
        price = sum(r["price"] * r["volume"] for r in group) / vol
        bids = sum(r["bids"] for r in group if r["bids"] is not None)
        cover = (
            bids / vol
            if bids
            else (group[0]["cover"] if len(group) == 1 else None)
        )
        out.append(
            {
                "date": day,
                "close": round(price, 3),
                "volume": int(vol),
                "total_bids": int(bids) if bids else "",
                "cover_ratio": round(cover, 3) if cover else "",
            }
        )
    return out


def import_archive(path: Path) -> list[dict]:
    rows: list[dict] = []
    if path.suffix == ".zip":
        with zipfile.ZipFile(path) as zf:
            for name in sorted(zf.namelist()):
                if name.endswith((".xls", ".xlsx")) and "auction-report" in name:
                    rows += parse_report(name, zf.read(name))
    else:
        for f in sorted(path.rglob("*.xls*")):
            rows += parse_report(f.name, f.read_bytes())
    return aggregate_daily(rows)


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: python -m eua_monitor.eex_import <zip-or-directory>", file=sys.stderr)
        return 2
    daily = import_archive(Path(sys.argv[1]))
    if not daily:
        print("error: no EUA auction rows found", file=sys.stderr)
        return 1
    config.EUA_HISTORY_CSV.parent.mkdir(parents=True, exist_ok=True)
    with config.EUA_HISTORY_CSV.open("w", newline="") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=["date", "close", "volume", "total_bids", "cover_ratio"]
        )
        writer.writeheader()
        writer.writerows(daily)
    print(
        f"wrote {config.EUA_HISTORY_CSV}: {len(daily)} auction days, "
        f"{daily[0]['date']} → {daily[-1]['date']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
